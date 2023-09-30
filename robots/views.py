from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from .models import Robot
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from .models import Robot
from django.db.models import Count


@method_decorator(csrf_exempt, name='dispatch')
class CreateRobotView(View):
    def post(self, request):
        try:
            data = json.loads(request.body.decode('utf-8'))
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')

            # Валидация данных
            if not model or not version or not created:
                return JsonResponse({'error': 'Все поля должны быть заполнены'}, status=400)

            # Создание записи о роботе
            robot = Robot(model=model, version=version, created=created)
            robot.save()

            return JsonResponse({'message': 'Запись о роботе успешно создана'}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Ошибка разбора JSON'}, status=400)


def excel_table_view(request):
    # Создаем новый Excel-файл
    workbook = Workbook()

    # Получаем текущую дату и вычисляем начальную и конечную даты для последней недели
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)

    # Получаем список уникальных моделей роботов
    unique_models = Robot.objects.values('model').distinct()

    # Перебираем модели и создаем листы для каждой модели
    for model in unique_models:
        model_name = model['model']
        sheet = workbook.create_sheet(title=model_name)  # Создаем лист с именем модели

        # Выполняем агрегацию для подсчета количества роботов каждой версии данной модели
        robots_summary = (
            Robot.objects
            .filter(model=model_name, created__gte=start_date, created__lte=end_date)
            .values('version')
            .annotate(quantity=Count('id'))
        )

        # Заголовки для столбцов
        headers = ['Модель', 'Версия', 'Количество за неделю']
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = header_title

        # Заполняем таблицу данными
        row_num = 2
        for summary in robots_summary:
            sheet[f'A{row_num}'] = model_name
            sheet[f'B{row_num}'] = summary['version']
            sheet[f'C{row_num}'] = summary['quantity']
            row_num += 1

    # Удаляем лист по умолчанию
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # Создаем HTTP-ответ для отображения Excel-файла
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=summary_report.xlsx'
    workbook.save(response)

    return response


def hello_world(request):
    return HttpResponse("Привет, мир!")
